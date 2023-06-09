import os
import re
import sys
import tkinter as tk
import webbrowser
from datetime import datetime
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk

import pandas as pd
import unicodedata
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

current_version = '0.30 (2023-04-28)'

# Set Pandas display options
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.expand_frame_repr', False)


# File handling

regex_pattern = r"(<.+?>)|(%[sdmyY])|({\d})|\((\+{\d})\)|({[A-Z]})|(\[[^\[]+\])|(\(\+\[[^\]]+\]\)%?)|(\d+\.?\d*%)|(" \
                r"\\n)|(\$\[[\w]+\])|(\{[A-Z_#0-9]+\})|(\bhttps?://\S+)|(\${\w+})|(&lt;t class=\"t_lc\"&gt;)|(" \
                r"&lt;/t&gt;)|@"

script_name = 'Translation Report Tool GI v.' + current_version

def get_xlsx_file_paths_in_folder(folder_path):
    """
    Returns a list of Excel file paths in a folder with an '.xlsx' extension.

    Args:
        folder_path (str): The path of the folder to search.

    Returns:
        list: A list of Excel file paths within the specified folder with an '.xlsx' extension.
    """
    try:
        xlsx_file_paths = []
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)
            if os.path.isfile(file_path) and file_name.lower().endswith('.xlsx'):
                xlsx_file_paths.append(file_path)
        return xlsx_file_paths
    except FileNotFoundError:
        print(f"Folder not found: {folder_path}")
        return []


# Chinese chars calc

def count_chinese_characters(s):
    """
    Counts the number of Chinese characters or words in a string.

    Args:
        s (str): The string to count Chinese characters or words in.

    Returns:
        int: The number of Chinese characters or words in the string.
    """
    s = str(s)
    if s is None:
        return 0

    global cjk_or_words_count  # this is to turn Chinese or Word calculation on or off

    if cjk_or_words_count == 'Words':
        count = len(s.split())
        return count
    else:
        count = 0
        for c in s:
            unicode_name = unicodedata.name(c, '')
            unicode_codepoint = ord(c)

            # Check for CJK Unified Ideographs (used in Chinese, Japanese, and Korean)
            if 'CJK UNIFIED IDEOGRAPH' in unicode_name:
                count += 1

            # Check for Hiragana (used in Japanese)
            elif 'HIRAGANA' in unicode_name:
                count += 1

            # Check for Katakana (used in Japanese)
            elif 'KATAKANA' in unicode_name:
                count += 1

            # Check for Hangul Syllables (used in Korean)
            elif 'HANGUL SYLLABLE' in unicode_name:
                count += 1

            # Check for CJK Symbols and Punctuation
            elif 0x3000 <= unicode_codepoint <= 0x303F:
                count += 1

            # Check for Halfwidth and Fullwidth Forms
            elif 0xFF00 <= unicode_codepoint <= 0xFFEF:
                count += 1
        return count


def count_regex(input_string):
    if not isinstance(input_string, str):
        return 0
    global regex_pattern
    regex = re.compile(regex_pattern)
    return len(regex.findall(input_string))


# Create a new DataFrame with headers from the report_headers argument
def create_report_dataframe(report_headers):
    # Create an empty DataFrame with the specified headers
    df = pd.DataFrame(columns=report_headers)
    return df


# Turn the whole Excel into a dict of dfs (this is faster than opening one by one)
def load_sheets_as_dict(excel_file, source_lang, target_lang):
    # Read the entire Excel file into memory
    all_sheets = pd.read_excel(excel_file, engine='openpyxl', sheet_name=None)

    # Initialize an empty dictionary to store sheet data
    sheets_data = {}

    # Iterate through sheet names and filter the specified columns
    for sheet_name, data in all_sheets.items():
        #        print(sheet_name)
        data = data[[source_lang, target_lang]]
        sheets_data[sheet_name] = data

    return sheets_data


# count chinese characters in column
def count_characters_in_column(df, column_name, count_function):
    # Apply the counting function to each value in the specified column
    character_counts = df[column_name].apply(count_function)

    # Calculate the total count of characters in the column
    total_characters = character_counts.sum()

    return total_characters


def count_regex_in_column(df, column_name, count_function):
    regex_count = df[column_name].apply(count_function)
    total_regex = regex_count.sum()
    return total_regex


def remove_empty_rows(df, target_column):
    # Remove rows where the target column is empty (NaN)
    filtered_df = df.dropna(subset=[target_column])

    return filtered_df


def count_unique_untranslated(df, target_column, source_column, count_function):
    column_df = df.copy()
    column_df = column_df.dropna(subset=[target_column])
    column_df.drop_duplicates(inplace=True)
    total_characters = column_df[source_column].apply(count_function).sum()
    return total_characters


def count_unique_characters(df, column_name, count_function):
    # Create a new dataframe with only the specified column
    column_df = df[[column_name]].copy()

    # Remove duplicate rows
    column_df.drop_duplicates(inplace=True)

    # Apply the count function to the content of the column and sum the results
    total_characters = column_df[column_name].apply(count_function).sum()

    return total_characters


# take file path and parameters, return a df to concatenate into the report df
def process_excel_file(excel_file, source_lang, target_lang, report_headers):
    filename_with_extension = os.path.basename(excel_file)
    # Read the entire Excel file into memory
    all_sheets = pd.read_excel(excel_file, engine='openpyxl', sheet_name=None)

    # Initialize an empty DataFrame for storing the processed data
    interim_df = pd.DataFrame(columns=report_headers)

    # Iterate through sheet names and process the data
    for sheet_name, data in all_sheets.items():
        # Filter the specified columns
        data = data[[source_lang, target_lang]]
        unique = count_unique_characters(data, source_lang, count_chinese_characters)
        source_chars = count_characters_in_column(data, source_lang, count_chinese_characters)
        regex_number = count_regex_in_column(data, source_lang, count_regex)
        data2 = data.copy()
        data = remove_empty_rows(data, target_lang)

        global selection_unique_or_all
        if selection_unique_or_all == 'All strings':
            translated_chars = count_characters_in_column(data, source_lang, count_chinese_characters)
            untranslated_chars = source_chars - translated_chars
        else:
            translated_chars = count_unique_untranslated(data2, target_lang, source_lang, count_chinese_characters)
            untranslated_chars = unique - translated_chars

        # need to update translated and untranslated functions to work properly!
        if selection_unique_or_all == "All strings":
            source_chars = source_chars
        else:
            source_chars = unique

        # translated_chars = count_characters_in_column(data, source_lang, count_chinese_characters)
        # untranslated_chars = source_chars - translated_chars
        if source_chars > 0:
            completeness = int((translated_chars / source_chars) * 100)
            code_and_variables_perc = int((regex_number / source_chars) * 100)
        else:
            completeness = 0
            code_and_variables_perc = 0

        # Create a new DataFrame with the data for this iteration
        row_data = pd.DataFrame({"Key": [sheet_name],
                                 "Source Wordcount": [source_chars],
                                 "Translated": [translated_chars],
                                 "Not_translated": [untranslated_chars],
                                 "file": [filename_with_extension],
                                 "Completeness": [completeness],
                                 "Variables ratio": [code_and_variables_perc],
                                 "Source Unique": [unique]})

        # Append the row_data to the interim_df
        interim_df = pd.concat([interim_df, row_data], ignore_index=True)

    return interim_df


def process_list_of_excels(report_dataframe, file_list, source_lang, target_lang, report_headers):
    for file_path in file_list:

        # Process the current Excel file
        current_result = process_excel_file(file_path, source_lang, target_lang, report_headers)

        # Check if the headers match
        if not set(current_result.columns) == set(report_dataframe.columns):
            # Add missing columns to report_dataframe with empty values
            for column in report_headers:
                if column not in current_result.columns:
                    current_result[column] = ""

        # Append the results to the report_dataframe
        report_dataframe = pd.concat([report_dataframe, current_result], ignore_index=True)

    # Calculate the sum of the relevant columns
    sum_row = report_dataframe[['Source Wordcount', 'Translated', 'Not_translated']].sum().to_frame().T
    sum_row['file'] = 'Total'
    sum_row['Key'] = '-'

    # Append the sum row to the bottom of the dataframe
    report_dataframe = pd.concat([report_dataframe, sum_row], ignore_index=True)

    return report_dataframe


def format_and_save_to_excel(df, filepath):
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Convert the DataFrame to rows and write them to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row):
            # Write the cell value
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)

            # Format the header row
            if r_idx == 0:
                cell.font = Font(bold=True)
                ws.freeze_panes = ws.cell(row=2, column=1)  # Freeze the header row
                ws.row_dimensions[1].height = 30  # Set header row height to 30 (double the default height)

            # Format the 'Completeness' column
            if ws.cell(row=1, column=c_idx + 1).value == 'Completeness' and r_idx > 0:  # Exclude header row
                value_float = float(value)  # Convert the cell value to a float
                if value_float == 100:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif value_float == 0:
                    cell.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                # elif 1 <= value_float <= 49:
                # cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.value = f'{value}%'
            # for regex marking
            if ws.cell(row=1, column=c_idx + 1).value == 'Variables ratio' and r_idx > 0:  # Exclude header row
                value = ws.cell(row=r_idx + 1, column=c_idx + 1).value
                if isinstance(value, int):
                    value_float = float(value)
                else:
                    value_float = value
                if value_float > 5:
                    cell.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                cell.value = f'{value_float}%'

            # Apply cell alignment
            alignment = Alignment(horizontal='center', vertical='center')
            cell.alignment = alignment

            # Apply word wrapping to the first two columns and the header row
            if c_idx < 2 or r_idx == 0:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    column_file = None
    for c_idx, column_title in enumerate(df.columns):
        if column_title == 'file':
            column_file = get_column_letter(c_idx + 1)
            break

    if column_file:
        start_merge = 2
        for row_idx in range(2, ws.max_row):
            if row_idx == ws.max_row - 1:  # Check if this is the second last row
                ws.merge_cells(f"{column_file}{start_merge}:{column_file}{row_idx}")
                break
            if ws[f"{column_file}{row_idx}"].value == ws[f"{column_file}{start_merge}"].value:
                continue
            else:
                ws.merge_cells(f"{column_file}{start_merge}:{column_file}{row_idx - 1}")
                start_merge = row_idx

    # Set column widths
    ws.column_dimensions['A'].width = 40  # 'file' column
    ws.column_dimensions['B'].width = 40  # 'Key' column
    for col_idx in range(2, len(df.columns)):
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 15

    # Save the workbook to the specified filepath
    wb.save(filepath)


# Variables that shouldn't be changed
report_headers_variable = [
    "file",
    "Key",
    "Source Wordcount",
    "Translated",
    "Not_translated",
    "Completeness",
    "Translator",
    "Proofreader",
    "Batch 1",
    "Batch 2",
    "Batch 3",
    "Batch 4",
    "Batch 5",
    "Batch 6",
    "Live",
    "Variables ratio",
    "Source Unique"
]
language_codes = ['RU', 'ES', 'FR', 'ID', 'JP', 'KR', 'PT', 'RU', 'TH', 'VI', 'TR', 'IT', 'CHS', 'en', 'kr', 'cht',
                  'jp', 'th', 'vi', 'id', 'es', 'ru', 'pt', 'de', 'fr', 'CHT', 'DE', 'EN',
                  'chs']
now = datetime.now()
timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")

# Variables that should be changed
folder_location = os.getcwd() + r'\source'

source_lang_codes_all = language_codes
# report_save_path = r'c:\2\report.xlsx'
report_save_path = os.getcwd() + (str(r'\reports\report_{}.xlsx'.format(timestamp)))

filelist = get_xlsx_file_paths_in_folder(folder_location)
report_df = create_report_dataframe(report_headers_variable)
output_filepath = ''


def read_and_save(df, wordlist, source, target, headers, output):
    one_file_df = process_list_of_excels(df, wordlist, source, target, headers)
    format_and_save_to_excel(one_file_df, output)


def for_button():
    try:
        read_and_save(report_df, filelist, source_lang_code.get(), target_lang_code.get(), report_headers_variable,
                      report_save_path)
        # Show popup window with message "Process complete"
        messagebox.showinfo("Process complete",
                            "Report has been generated. You can find it at: " + str(report_save_path))
    except Exception as e:
        # Show popup window with error message
        messagebox.showerror("Error", str(e))
    print("Button clicked")


def browse_folder():
    global folder_location
    folder_location = filedialog.askdirectory()
    folder_path_var.set(folder_location)
    print(folder_location)
    global filelist
    filelist = get_xlsx_file_paths_in_folder(folder_location)


def save_report():
    global report_save_path
    appendix = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    report_save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=f"report_{appendix}")
    report_save_path_label.config(text=report_save_path)


# Create a new window object
window = tk.Tk()
window_name = script_name
# Set the window title
window.title(window_name)

# Set the window size
window.geometry("620x400")

# Create a frame to hold the browse button and file path
frame = tk.Frame(window)

# Info text
info_text = tk.Label(window, text="Select a folder with source *.xlsx files")
info_text.grid(row=0, column=0, sticky='w', padx=10, pady=0)

# Create a frame to hold the browse button and file path
frame.grid(row=1, column=0, padx=10, pady=10, sticky='w')

# Create a button to browse for a folder
folder_path_var = tk.StringVar()
browse_button = ttk.Button(frame, text="Browse folder", command=browse_folder)
browse_button.grid(row=1, column=0, padx=0, pady=5, sticky='w')

# Create a text field to display the folder path

folder_path_var.set(folder_location)
folder_path_entry = tk.Entry(frame, textvariable=folder_path_var, width=80)
folder_path_entry.grid(row=1, column=0, padx=120, pady=5, sticky='w')

# Elements for saving report
save_report_button = ttk.Button(window, text="Save report to...", command=save_report)
save_report_button.grid(row=2, column=0, padx=10, pady=10, sticky='w')
report_save_path_label = tk.Label(window, text=report_save_path)
report_save_path_label.grid(row=2, column=0, padx=130, pady=10, sticky='w')

# Elements for language codes
lang_codes_label1 = tk.Label(window, text="Source Language Code:")
lang_codes_label1.grid(row=3, column=0, sticky='w', padx=10, pady=10)
source_lang_code = tk.StringVar()
source_lang_combobox = ttk.Combobox(window, textvariable=source_lang_code, values=source_lang_codes_all, width=5,
                                    state='readonly')
source_lang_combobox.current(source_lang_codes_all.index('CHS'))
source_lang_combobox.grid(row=3, column=0, sticky='w', padx=150, pady=10)

lang_codes_label2 = tk.Label(window, text="Target Language:")
lang_codes_label2.grid(row=6, column=0, sticky='w', padx=10, pady=10)

target_lang_code = tk.StringVar()
target_lang_combobox = ttk.Combobox(window, textvariable=target_lang_code, values=source_lang_codes_all, width=5,
                                    state='readonly')
target_lang_combobox.current(source_lang_codes_all.index('RU'))
target_lang_combobox.grid(row=6, column=0, sticky='w', padx=150, pady=10)

# Button to process files
process_button = ttk.Button(window, text="Generate report", command=for_button)
process_button.grid(row=9, column=0, padx=10, pady=10, sticky='w')


# unique change

def on_unique_or_all_change(*args):
    global selection_unique_or_all
    selection_unique_or_all = unique_or_all_var.get()
    print("selection_unique_or_all:", selection_unique_or_all)


# Create a StringVar to hold the selected option
unique_or_all_var = tk.StringVar(window)
unique_or_all_var.trace('w', on_unique_or_all_change)

# Set the default value
unique_or_all_var.set("All strings")
selection_unique_or_all = unique_or_all_var.get()

# Create the dropdown menu

# Create a custom style for the dropdown menu
options = ["All strings", "Unique only"]
dropdown2 = tk.OptionMenu(window, unique_or_all_var, *options)
source_count_label2 = tk.Label(window, text="Count all strings or unique only?")
source_count_label2.grid(row=8, column=0, sticky='w', padx=10, pady=10)
dropdown2.grid(row=8, column=0, padx=200, pady=5, sticky='w')


# count source change

def on_option_change(*args):
    global cjk_or_words_count
    cjk_or_words_count = var.get()
    print("cjk_or_words_count:", cjk_or_words_count)


# Create a StringVar to hold the selected option
var = tk.StringVar(window)
var.trace('w', on_option_change)

var.set("Chinese")
cjk_or_words_count = var.get()

# Create a custom style for the dropdown menu


options = ["Chinese", "Words"]
dropdown = tk.OptionMenu(window, var, *options)
source_count_label = tk.Label(window, text="Count source Chinese or Words?")
source_count_label.grid(row=7, column=0, sticky='w', padx=10, pady=10)
dropdown.grid(row=7, column=0, padx=200, pady=10, sticky='w')


# Text in the bottom
def open_url(url):
    webbrowser.open(url)


about_label = tk.Label(window, text="github.com/wtigga\nVladimir Zhdanov", fg="blue", cursor="hand2", justify="left")
about_text = tk.Label(window, text=current_version)
about_text.grid(row=10, column=0, sticky='w', padx=10, pady=0)
about_label.bind("<Button-1>",
                 lambda event: open_url("https://github.com/wtigga/TranslationSourceExcelFilesReportGeneratorGI"))
about_label.grid(row=11, column=0, sticky='w', padx=10, pady=0)


# console output
class TextRedirector:
    def __init__(self, widget):
        self.widget = widget

    def write(self, text):
        self.widget.configure(state='normal')
        self.widget.insert(tk.END, text)
        self.widget.see(tk.END)
        self.widget.configure(state='disabled')

    def flush(self):
        pass


output_text = tk.Text(window, wrap='word', height=10, state='disabled')
output_text.grid(row=12, column=0, sticky='nsew')

sys.stdout = TextRedirector(output_text)


# tooltips
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None

        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)

    def on_enter(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()

        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25

        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")

        label = tk.Label(self.tip_window, text=self.text, background="#ffffe0", relief="solid", borderwidth=1)
        label.pack()

    def on_leave(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


source_count_label_tooltip_text = 'Chinese is suitable when source is in Chinese or Japanese. It will count only ' \
                                  'those characters.\nWords suitable for other languages, like English, it will count' \
                                  ' words (delimited by a space).'
source_count_label_tooltip = ToolTip(source_count_label, source_count_label_tooltip_text)

source_count_label2_tooltip_text = 'All strings will count as is. Unique only will first drop the 100% duplicates in ' \
                                   'source in each sheet (ID).\nMight be a slight inconsistency in completeness.'
source_count_label2_tooltip = ToolTip(source_count_label2, source_count_label2_tooltip_text)

lang_codes_label1_tooltip_text = 'Case sensitive - "EN" and "en" are not the same.'
lang_codes_label1_tooltip = ToolTip(lang_codes_label1, lang_codes_label1_tooltip_text)
lang_codes_label2_tooltip = ToolTip(lang_codes_label2, lang_codes_label1_tooltip_text)

info_text_tooltip_text = 'The folder must only contain source files, and nothing else.'
info_text_tooltip = ToolTip(info_text, info_text_tooltip_text)
browse_button_tooltip = ToolTip(browse_button, info_text_tooltip_text)

# Start the main event loop
window.mainloop()

'''While the logic and architecture are products of the author's thinking capabilities,
lots of functions in the code were written with the help of OpenAi's ChatGPT 3.5 and ChatGPT 4.'''
